"""Excel/Image → CPT 分析器

用法:
    python analyze_upload.py sales_report.xlsx
    python analyze_upload.py table_photo.png
    python analyze_upload.py scanned_form.jpg
"""

import sys, os, json, re, base64
from openpyxl import load_workbook

OUT_DIR = r'os.path.dirname(os.path.abspath(__file__))\output'
os.makedirs(OUT_DIR, exist_ok=True)

# ============== Excel 解析 ==============

def parse_excel(filepath):
    """解析 Excel，提取表头+样本数据"""
    wb = load_workbook(filepath, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return None
    
    # 找表头行（第一个所有cell都非空的行）
    header_row_idx = 0
    for i, row in enumerate(rows[:10]):
        non_empty = sum(1 for c in row if c is not None and str(c).strip())
        if non_empty >= 2:
            header_row_idx = i
            break
    
    headers = [str(c).strip() if c else '' for c in rows[header_row_idx]]
    headers = [h for h in headers if h]
    
    # 取样本数据行（跳过表头）
    sample_rows = []
    for row in rows[header_row_idx + 1:]:
        filtered = [str(c).strip() if c is not None else '' for c in row[:len(headers)]]
        if any(filtered):
            sample_rows.append(filtered)
        if len(sample_rows) >= 5:
            break
    
    # 推断字段类型
    field_types = []
    for i in range(len(headers)):
        vals = [r[i] for r in sample_rows if i < len(r) and r[i]]
        if all(re.match(r'^\d+(\.\d+)?$', v) for v in vals if v):
            field_types.append('NUMBER')
        elif all(re.match(r'^\d{4}[-/]\d{1,2}[-/]\d{1,2}', v) for v in vals if v):
            field_types.append('DATE')
        else:
            field_types.append('VARCHAR')
    
    return {
        'headers': headers,
        'data_types': field_types,
        'sample_data': sample_rows[:3],
        'row_count': len(list(ws.iter_rows())) - header_row_idx - 1,
        'sheet_name': ws.title
    }


# ============== 图片解析 ==============

def parse_image(filepath):
    """解析图片中的表格，使用 OCR + LLM"""
    # Step 1: OCR
    import pytesseract
    from PIL import Image
    
    pytesseract.pytesseract.tesseract_cmd = r'C:\Program Files\Tesseract-OCR\tesseract.exe'
    
    img = Image.open(filepath)
    # 尝试中英文混合OCR
    try:
        text = pytesseract.image_to_string(img, lang='chi_sim+eng')
    except:
        text = pytesseract.image_to_string(img, lang='eng')
    
    if not text.strip():
        # 尝试用 PIL 分析颜色布局来检测表格
        text = _detect_table_from_image(filepath)
    
    return {'ocr_text': text.strip(), 'source': 'ocr'}


def _detect_table_from_image(filepath):
    """从图片像素检测表格结构（fallback）"""
    from PIL import Image
    import numpy as np
    img = Image.open(filepath).convert('L')
    arr = np.array(img)
    # 简单边缘检测找表格线
    return f"[图片表格 - 需要LLM分析] 图片尺寸: {img.size}"


# ============== LLM 结构化提取 ==============

def llm_extract_columns(data):
    """用 LLM 从解析结果中提取结构化列定义"""
    from openai import OpenAI
    c = OpenAI(api_key='sk-79778f1a65f1484f81e863beb2ade2ee', base_url='https://api.deepseek.com')
    
    if 'headers' in data:
        # Excel 模式
        prompt = f"""从Excel表格解析结果提取FineReport列定义。输出JSON:
{{"title":"报表标题","columns":[{{"name":"中文列名","field":"英文字段名","type":"VARCHAR/NUMBER/DATE"}}],"datasource":"ds1","table":"表名"}}

Excel表头: {data['headers']}
数据类型: {data['types']}
样本数据: {data['sample'][:3]}
"""
    else:
        # 图片OCR模式
        prompt = f"""从OCR识别的表格文本中提取FineReport列定义。输出JSON:
{{"title":"报表标题","columns":[{{"name":"中文列名","field":"英文字段名"}}],"datasource":"ds1"}}

OCR文本:
{data['ocr_text'][:2000]}
"""
    
    try:
        r = c.chat.completions.create(
            model='deepseek-chat',
            messages=[{'role':'user','content':prompt}],
            temperature=0, max_tokens=600, timeout=15
        )
        raw = r.choices[0].message.content
        m = re.search(r'\{.*\}', raw, re.DOTALL)
        if m:
            return json.loads(m.group())
    except Exception as e:
        print(f'  LLM error: {e}')
    
    return None


# ============== 主入口 ==============

def analyze(filepath):
    """主入口：分析文件 → 输出 CPT"""
    ext = os.path.splitext(filepath)[1].lower()
    print(f'Analyzing: {filepath}')
    
    # Step 1: 解析文件
    if ext in ('.xlsx', '.xls'):
        print('  Mode: Excel')
        data = parse_excel(filepath)
        if data is None:
            print('  ERROR: Cannot parse Excel')
            return
        print(f'  Sheet: {data["sheet_name"]}')
        print(f'  Headers ({len(data["headers"])}): {data["headers"][:10]}')
        print(f'  Sample: {data["sample_data"][:2]}')
        
        llm_input = {
            'headers': data['headers'],
            'types': data['data_types'],
            'sample': data['sample_data']
        }
        
    elif ext in ('.png', '.jpg', '.jpeg'):
        print('  Mode: Image/OCR')
        data = parse_image(filepath)
        print(f'  OCR text ({len(data["ocr_text"])} chars):')
        print(data['ocr_text'][:300])
        
        llm_input = {'ocr_text': data['ocr_text']}
        
    else:
        print(f'  ERROR: Unsupported format {ext}')
        return
    
    # Step 2: LLM 提取结构化列
    print('\n  Extracting columns with LLM...')
    params = llm_extract_columns(llm_input)
    if not params:
        print('  ERROR: LLM extraction failed')
        return
    
    title = params.get('title', '报表')
    columns = params.get('columns', [])
    print(f'  Title: {title}')
    print(f'  Columns: {len(columns)}')
    for col in columns:
        print(f'    {col["name"]} -> {col["field"]}')
    
    # Step 3: 生成 CPT
    # 检查真实数据库中是否有匹配的表
    real_table = None
    try:
        with open(r'C:\workspace\01_knowledge\parsed\datasource_catalog.json', encoding='utf-8') as f:
            cat = json.load(f)
        # 这里应该查真实数据库，但先跳过
    except:
        pass
    
    from cpt_builder_v4 import build_v10_cpt
    fields = [c['field'] for c in columns]
    table = params.get('table', 'report_data')
    
    sql = f"SELECT {', '.join(fields)} FROM [{table}] ORDER BY 1"
    
    cpt_xml = build_v10_cpt(
        title=title,
        columns=columns,
        table_name=table,
        sql=sql,
        db_name='FRDemo'
    )
    
    # Step 4: 保存
    safe_title = re.sub(r'[^\w\u4e00-\u9fff]', '_', title)[:30]
    cpt_path = os.path.join(OUT_DIR, f'{safe_title}.cpt')
    with open(cpt_path, 'w', encoding='utf-8') as f:
        f.write(cpt_xml)
    
    print(f'\n  CPT saved: {cpt_path} ({len(cpt_xml)} bytes)')
    print(f'  SQL: {sql}')
    
    # 也部署到 FineReport
    import shutil
    fr_path = r'/path/to/fr/reportlets'
    shutil.copy2(cpt_path, os.path.join(fr_path, f'{safe_title}.cpt'))
    print(f'  Deployed to FineReport reportlets')
    
    return {'cpt_path': cpt_path, 'params': params}


if __name__ == '__main__':
    if len(sys.argv) > 1:
        analyze(sys.argv[1])
    else:
        print('Usage: python analyze_upload.py <file.xlsx|file.png|file.jpg>')
